#!/usr/bin/env python3
import json
import gzip
import os
import re
import sys
import argparse
import requests
from datetime import datetime

AA3TO1 = {
    "Ala": "A", "Arg": "R", "Asn": "N", "Asp": "D", "Cys": "C", "Gln": "Q", "Glu": "E", "Gly": "G",
    "His": "H", "Ile": "I", "Leu": "L", "Lys": "K", "Met": "M", "Phe": "F", "Pro": "P", "Ser": "S",
    "Thr": "T", "Trp": "W", "Tyr": "Y", "Val": "V", "Ter": "*"
}

GENE_LIST = set()
CODING_CSQ = set()
SPLICE_CSQ = set()
ALWAYS_EXCLUDED_CSQ = set()
TERT_HOTSPOTS = {}

ONCOKB_TOKEN = os.environ.get("ONCOKB_API_TOKEN", "")
ONCOKB_BASE = "https://www.oncokb.org/api/v1"

# --- Helper Functions ---

def load_json_file(file_path):
    if file_path.endswith(".gz"):
        with gzip.open(file_path, 'rt', encoding='utf-8') as f:
            return json.load(f)
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def detect_vcs_label(data):
    if not isinstance(data, dict):
        return None

    for key_path in (
        ("vcs",),
        ("vcsType",),
        ("sciobaseVcs",),
        ("metadata", "vcs"),
        ("metadata", "vcsType"),
        ("metadata", "sciobaseVcs"),
    ):
        node = data
        for key in key_path:
            if not isinstance(node, dict) or key not in node:
                node = None
                break
            node = node.get(key)
        if node:
            return str(node)

    return None

def load_filter_config(config_path, vcs_label=None):
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Filter config not found at {config_path}")

    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)

    profiles = config.get("profiles") if isinstance(config, dict) else None
    if profiles is None and isinstance(config, dict):
        profiles = config
    if not isinstance(profiles, dict):
        profiles = {}

    default_name = config.get("default_vcs") if isinstance(config, dict) else None
    selected_name = vcs_label or default_name or "default"

    profile = profiles.get(selected_name) or profiles.get("default") or {}
    if not profile and isinstance(config, dict):
        profile = config

    merged = {}
    for key in ("gene_list", "coding_csq", "splice_csq", "always_excluded_csq"):
        value = profile.get(key)
        if value:
            merged[key] = value

    merged["min_depth"] = profile.get("min_depth", 10)
    merged["max_popmax"] = profile.get("max_popmax", 0.01)
    merged["main_min_vaf"] = profile.get("main_min_vaf", 0.05)

    tert_hotspots = profile.get("tert_hotspots")
    if isinstance(tert_hotspots, dict) and tert_hotspots:
        merged["tert_hotspots"] = tert_hotspots

    return selected_name, merged

def apply_filter_profile(profile):
    global GENE_LIST, CODING_CSQ, SPLICE_CSQ, ALWAYS_EXCLUDED_CSQ, TERT_HOTSPOTS

    GENE_LIST = set(profile.get("gene_list", []))
    CODING_CSQ = set(profile.get("coding_csq", []))
    SPLICE_CSQ = set(profile.get("splice_csq", []))
    ALWAYS_EXCLUDED_CSQ = set(profile.get("always_excluded_csq", []))
    TERT_HOTSPOTS = dict(profile.get("tert_hotspots", {}))

def get_popmax(v):
    af_values = []
    if "gnomad" in v and v["gnomad"].get("allAf") is not None:
        af_values.append(v["gnomad"]["allAf"])
    if "oneKg" in v and v["oneKg"].get("allAf") is not None:
        af_values.append(v["oneKg"]["allAf"])
    if "esp" in v and v["esp"].get("allAf") is not None:
        af_values.append(v["esp"]["allAf"])
    return max(af_values) if af_values else None

def convert_aa_3to1(hgvsp):
    if not hgvsp:
        return hgvsp
    res = hgvsp
    for aa3, aa1 in AA3TO1.items():
        res = res.replace(aa3, aa1)
    return res

def build_hgvs(tx, hgvsg):
    hgvsc = tx.get("hgvsc", "")
    hgvsp = tx.get("hgvsp", "")
    p_match = re.search(r"(p\.\S+)", hgvsp)
    p_part = p_match.group(1) if p_match else ""
    
    if p_part:
        p_part = convert_aa_3to1(p_part)
        
    if hgvsc and p_part:
        return f"{hgvsc};{p_part}"
    if hgvsc:
        return hgvsc
    if p_part:
        return p_part
    return hgvsg or ""

def format_hgvs_genomic(vid, hgvsg=None):
    if vid:
        m = re.match(r"^(?:chr)?(\d+|[XYM])-(\d+)-([A-Z]+)-([A-Z]+)$", vid, re.IGNORECASE)
        if m:
            chrom, pos, ref, alt = m.groups()
            return f"{chrom}:g.{pos}{ref}>{alt}"
    
    if hgvsg:
        m = re.match(r"^NC_0*(\d+)\.\d+:g\.(\d+)([A-Z]+)>([A-Z]+)$", hgvsg)
        if m:
            chrom_num, pos, ref, alt = m.groups()
            chrom = str(int(chrom_num))
            if chrom == "23": chrom = "X"
            elif chrom == "24": chrom = "Y"
            return f"{chrom}:g.{pos}{ref}>{alt}"
            
        m = re.match(r"^(?:chr)?(\d+|[XYM]):g\.(\d+)([A-Z]+)>([A-Z]+)$", hgvsg, re.IGNORECASE)
        if m:
            chrom, pos, ref, alt = m.groups()
            return f"{chrom}:g.{pos}{ref}>{alt}"
            
    return hgvsg or ""

def extract_alteration(hgvsp):
    if not hgvsp:
        return None
    p_match = re.search(r"(p\.\S+)", hgvsp)
    if not p_match:
        return None
    p_str = p_match.group(1)
    # fs
    fs_m = re.match(r"p\.\(?([A-Z][a-z]{2})(\d+)([A-Z][a-z]{2})fs(?:Ter)?(\d+)?\)?", p_str)
    if fs_m:
        aa1 = AA3TO1.get(fs_m.group(1), fs_m.group(1))
        pos = fs_m.group(2)
        aa2 = AA3TO1.get(fs_m.group(3), fs_m.group(3))
        term = fs_m.group(4) or ""
        return f"{aa1}{pos}{aa2}fs*{term}"
    # stop gained
    sg_m = re.match(r"p\.\(?([A-Z][a-z]{2})(\d+)(Ter)\)?", p_str)
    if sg_m:
        aa1 = AA3TO1.get(sg_m.group(1), sg_m.group(1))
        pos = sg_m.group(2)
        return f"{aa1}{pos}*"
    # missense
    ms_m = re.match(r"p\.\(?([A-Z][a-z]{2})(\d+)([A-Z][a-z]{2})\)?", p_str)
    if ms_m:
        aa1 = AA3TO1.get(ms_m.group(1), ms_m.group(1))
        pos = ms_m.group(2)
        aa2 = AA3TO1.get(ms_m.group(3), ms_m.group(3))
        return f"{aa1}{pos}{aa2}"
    # del/dup/ins
    del_m = re.match(r"p\.\(?([A-Z][a-z]{2})(\d+)(del|dup|ins)\)?", p_str)
    if del_m:
        aa1 = AA3TO1.get(del_m.group(1), del_m.group(1))
        pos = del_m.group(2)
        variant_type = del_m.group(3)
        return f"{aa1}{pos}{variant_type}"
    return None

def categorize_variant(gene, consequences):
    csq_set = set(consequences)
    if csq_set:
        if all(c in ALWAYS_EXCLUDED_CSQ for c in csq_set):
            return "excluded"
    
    if any(c in CODING_CSQ for c in csq_set):
        return "coding"
    if gene == "TERT":
        return "tert_nc"
    if any(c in SPLICE_CSQ for c in csq_set):
        return "splice"
    
    return "excluded"

def map_classification(val):
    if not val:
        return ""
    return val.strip()

def classify_variant(v):
    if not ONCOKB_TOKEN:
        return None
    
    headers = {"Authorization": f"Bearer {ONCOKB_TOKEN}", "Accept": "application/json"}
    
    classification = ""
    
    # 1. Try genomic query (byHGVSg)
    if v.get("hgvsg"):
        try:
            r = requests.get(f"{ONCOKB_BASE}/annotate/mutations/byHGVSg?hgvsg={v['hgvsg']}&referenceGenome=GRCh37", 
                             headers=headers, timeout=10)
            if r.status_code == 200:
                data = r.json()
                if data.get("oncogenic"):
                    classification = map_classification(data["oncogenic"])
        except Exception:
            pass
            
    # 2. If no valid classification found, try fallback by protein change
    if not classification and v.get("alteration") and v.get("gene"):
        try:
            r = requests.get(f"{ONCOKB_BASE}/annotate/mutations/byProteinChange?hugoSymbol={v['gene']}&alteration={v['alteration']}&referenceGenome=GRCh37",
                             headers=headers, timeout=10)
            if r.status_code == 200:
                data = r.json()
                if data.get("oncogenic"):
                    classification = map_classification(data["oncogenic"])
        except Exception:
            pass
            
    return classification

# --- Main Parsing Logic ---

def load_variant_comments(reference_dir):
    comment_map = {}
    csv_path = os.path.join(reference_dir, "variant_comments.csv")
    if not os.path.exists(csv_path):
        # Look for variant_comments_*.csv
        files = [f for f in os.listdir(reference_dir) if f.startswith("variant_comments") and f.endswith(".csv")]
        if not files:
            print(f"Warning: No variant comments CSV found in {reference_dir}")
            return comment_map
        # Sort by date if possible, or just take the latest alphabetical
        files.sort()
        csv_path = os.path.join(reference_dir, files[-1])
        
    print(f"Loading variant comments from {csv_path}")
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            if not lines: return comment_map
            header = lines[0].strip().split("\t")
            try:
                chr_idx = header.index("Chromosome")
                pos_idx = header.index("Position")
                ref_idx = header.index("Ref_allele")
                alt_idx = header.index("Alt_allele")
                comment_idx = header.index("Comment")
                date_idx = header.index("Date") if "Date" in header else -1
                symbol_idx = header.index("Symbol") if "Symbol" in header else -1
                hgvsc_idx = header.index("HGVSC") if "HGVSC" in header else -1
            except ValueError as e:
                print(f"Error parsing header in {csv_path}: {e}")
                return comment_map

            for line in lines[1:]:
                cols = line.strip().split("\t")
                if len(cols) <= comment_idx: continue
                key = f"{cols[chr_idx]}-{cols[pos_idx]}-{cols[ref_idx]}-{cols[alt_idx]}"
                entry = {
                    "comment": cols[comment_idx].strip(),
                    "date": cols[date_idx].strip() if date_idx >= 0 and len(cols) > date_idx else "",
                    "gene": cols[symbol_idx].strip() if symbol_idx >= 0 and len(cols) > symbol_idx else "",
                    "hgvsc": cols[hgvsc_idx].strip() if hgvsc_idx >= 0 and len(cols) > hgvsc_idx else "",
                }
                if not entry["comment"]: continue
                if key not in comment_map:
                    comment_map[key] = []
                comment_map[key].append(entry)
    except Exception as e:
        print(f"Error reading variant comments: {e}")
    
    return comment_map

def get_comments_for_variant(vid, comment_map):
    if not vid or not comment_map:
        return []
    entries = comment_map.get(vid)
    if not entries:
        return []

    # Sort by date descending and remove duplicate comment rows while preserving order.
    sorted_entries = sorted(
        entries,
        key=lambda x: (x.get("date", ""), x.get("comment", ""), x.get("gene", ""), x.get("hgvsc", "")),
        reverse=True,
    )
    seen = set()
    unique_entries = []
    for entry in sorted_entries:
        key = (
            entry.get("date", ""),
            entry.get("comment", ""),
            entry.get("gene", ""),
            entry.get("hgvsc", ""),
        )
        if key in seen:
            continue
        seen.add(key)
        unique_entries.append(entry)
    return unique_entries

def process_annotation_file(file_path, comment_map, filter_profile=None):
    if filter_profile is None:
        raise ValueError("filter_profile is required")
    data = load_json_file(file_path)

    main_candidates = []
    qc_variants = []

    for pos in data.get("positions", []):
        if not pos.get("filters") or "PASS" not in pos["filters"] or len(pos["filters"]) != 1:
            continue
        
        samples = pos.get("samples", [])
        if not samples: continue
        sample = samples[0]
        depth = sample.get("totalDepth", 0)
        if depth < filter_profile.get("min_depth", 10): continue
        v_freqs = sample.get("variantFrequencies", [])

        for vi, variant in enumerate(pos.get("variants", [])):
            vaf = v_freqs[vi] if vi < len(v_freqs) else None
            pm = get_popmax(variant)

            vid = variant.get("vid", "")
            
            # Filter out non-coding variants, keeping only coding, splice, or TERT hotspot variants
            alt_allele = pos.get("altAlleles", [])[vi] if vi < len(pos.get("altAlleles", [])) else ""
            tert_key = f"{pos.get('chromosome')}-{pos.get('position')}-{pos.get('refAllele')}-{alt_allele}"
            is_tert_hs = tert_key in TERT_HOTSPOTS
            
            is_coding = False
            for tx in variant.get("transcripts", []):
                category = categorize_variant(tx.get("hgnc"), tx.get("consequence", []))
                if category != "excluded":
                    is_coding = True
                    break
            
            if not is_coding and not is_tert_hs:
                continue
            
            # QC Variant logic (PASS + depth >= 10 + popmax <= 1%)
            # Only keep TERT mutations, or coding variants with a 'p.' annotation
            selected_tx = None
            for tx in variant.get("transcripts", []):
                if tx.get("source") == "RefSeq" and tx.get("isCanonical"):
                    gene_name = tx.get("hgnc", "")
                    hgvsp_val = tx.get("hgvsp", "")
                    category = categorize_variant(gene_name, tx.get("consequence", []))
                    has_p_annotation = "p." in hgvsp_val
                    
                    if gene_name == "TERT" or (category == "coding" and has_p_annotation):
                        selected_tx = tx
                        break
            
            # Fallback to any transcript matching the rule if no canonical RefSeq transcript matches
            if not selected_tx:
                for tx in variant.get("transcripts", []):
                    gene_name = tx.get("hgnc", "")
                    hgvsp_val = tx.get("hgvsp", "")
                    category = categorize_variant(gene_name, tx.get("consequence", []))
                    has_p_annotation = "p." in hgvsp_val
                    
                    if gene_name == "TERT" or (category == "coding" and has_p_annotation):
                        selected_tx = tx
                        break
                        
            if selected_tx:
                qc_variants.append({
                    "gene": selected_tx.get("hgnc", ""),
                    "hgvs": build_hgvs(selected_tx, variant.get("hgvsg")),
                    "hgvsg": format_hgvs_genomic(vid, variant.get("hgvsg")),
                    "vid": vid,
                    "depth": depth,
                    "vaf": vaf
                })

            # Main Candidate logic (Gene list + consequences)
            for tx in variant.get("transcripts", []):
                gene_name = tx.get("hgnc", "")
                if tx.get("source") != "RefSeq" or not tx.get("isCanonical") or gene_name not in GENE_LIST:
                    continue
                
                category = categorize_variant(gene_name, tx.get("consequence", []))
                if category == "excluded": continue

                if pm is not None and pm > filter_profile.get("max_popmax", 0.01):
                    continue

                alt_allele = pos.get("altAlleles", [])[vi] if vi < len(pos.get("altAlleles", [])) else ""
                tert_key = f"{pos.get('chromosome')}-{pos.get('position')}-{pos.get('refAllele')}-{alt_allele}"
                tert_hs = TERT_HOTSPOTS.get(tert_key)

                candidate = {
                    "gene": tx.get("hgnc"),
                    "hgvs": tert_hs["label"] if tert_hs else build_hgvs(tx, variant.get("hgvsg")),
                    "hgvsg": format_hgvs_genomic(vid, variant.get("hgvsg")),
                    "alteration": extract_alteration(tx.get("hgvsp")),
                    "vid": vid,
                    "depth": depth,
                    "vaf": vaf,
                    "category": category,
                    "classification": tert_hs["classification"] if tert_hs else None,
                    "isTertHotspot": bool(tert_hs)
                }
                main_candidates.append(candidate)

    # Sort main candidates
    main_candidates.sort(key=lambda x: (x["gene"], -(x["vaf"] or 0)))

    # OncoKB Classification
    if main_candidates and ONCOKB_TOKEN:
        print(f"Classifying {len(main_candidates)} variants with OncoKB...")
        for v in main_candidates:
            if not v["isTertHotspot"] and not v["classification"]:
                v["classification"] = classify_variant(v)
    elif not ONCOKB_TOKEN:
        print("Warning: ONCOKB_API_TOKEN not set. Skipping OncoKB classification.")

    # Filter Main Variants (Splice must be pathogenic, VAF > 5%)
    main_variants = []
    for v in main_candidates:
        if v["category"] == "splice":
            if (v["classification"] or "").strip().lower() not in ["oncogenic", "likely oncogenic", "predicted oncogenic"]:
                continue
        if v["vaf"] is None or v["vaf"] <= filter_profile.get("main_min_vaf", 0.05):
            continue
        main_variants.append(v)

    # Mark variants as main candidates or background
    all_variants = []
    seen_vids = set()
    
    # Process Main Variants first
    for v in main_variants:
        v["is_main_candidate"] = True
        all_variants.append(v)
        seen_vids.add(v["vid"])
    
    # Process QC Variants second (add if not already in main)
    for v in qc_variants:
        if v["vid"] not in seen_vids:
            v["is_main_candidate"] = False
            all_variants.append(v)
            seen_vids.add(v["vid"])

    # Collect Comments
    def collect_comments(variants):
        comments = []
        seen = set()
        for v in variants:
            vid = v.get("vid")
            if not vid or vid in seen: continue
            c_list = get_comments_for_variant(vid, comment_map)
            if c_list:
                seen.add(vid)
                for c in c_list:
                    comments.append({
                        "gene": c.get("gene") or v.get("gene") or "",
                        "hgvsc": c.get("hgvsc", ""),
                        "comment": c.get("comment"),
                        "date": c.get("date"),
                        "vid": vid
                    })
        return comments

    all_comments = collect_comments(all_variants)

    return {
        "variants": all_variants,
        "comments": all_comments
    }

def write_variants_to_xlsx(results, xlsx_path):
    try:
        import pandas as pd
        from openpyxl.styles import PatternFill

        df_variants = pd.DataFrame(results.get("variants", []))

        # Track main candidate rows and drop the column
        main_candidate_rows = []
        if "is_main_candidate" in df_variants.columns:
            main_candidate_rows = df_variants.index[df_variants["is_main_candidate"] == True].tolist()
            df_variants = df_variants.drop(columns=["is_main_candidate"])

        # Reorder columns
        desired_order = ["gene", "hgvs", "classification", "depth", "vaf"]
        for col in desired_order:
            if col not in df_variants.columns:
                df_variants[col] = None
        
        remaining_cols = [c for c in df_variants.columns if c not in desired_order]
        df_variants = df_variants[desired_order + remaining_cols]

        df_comments = pd.DataFrame(results.get("comments", []))

        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            df_variants.to_excel(writer, sheet_name="Variants", index=False)
            df_comments.to_excel(writer, sheet_name="Comments", index=False)

            # Apply row highlighting: main candidates yellow, high VAF background variants red/pink
            worksheet_var = writer.sheets["Variants"]
            main_highlight = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            vaf_highlight = PatternFill(start_color="FFEAE6", end_color="FFEAE6", fill_type="solid")
            
            for row_idx in range(2, worksheet_var.max_row + 1):
                df_idx = row_idx - 2
                is_main = df_idx in main_candidate_rows
                
                # Check VAF from df_variants
                vaf_val = df_variants.iloc[df_idx].get("vaf")
                is_high_vaf = vaf_val is not None and vaf_val > 0.05
                
                if is_main:
                    for col_idx in range(1, worksheet_var.max_column + 1):
                        worksheet_var.cell(row=row_idx, column=col_idx).fill = main_highlight
                elif is_high_vaf:
                    for col_idx in range(1, worksheet_var.max_column + 1):
                        worksheet_var.cell(row=row_idx, column=col_idx).fill = vaf_highlight

            # Auto-fit column widths
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for col in worksheet.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        val_to_check = str(cell.value or '')
                        if cell.value is not None:
                            max_len = max(max_len, max(len(line) for line in val_to_check.split('\n')))
                    worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)
        print(f"Processed variants Excel saved to {xlsx_path}")
    except Exception as e:
        print(f"Error saving variants Excel: {e}", file=sys.stderr)

def main():
    parser = argparse.ArgumentParser(description="Extract and filter NGS variants for report.")
    parser.add_argument("input_json", help="Path to input VEP JSON(.gz) file.")
    default_ref_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "resources")
    default_filter_config = os.path.join(default_ref_dir, "sciobase_filters.json")
    parser.add_argument("--ref-dir", default=default_ref_dir, help="Directory containing variant_comments.csv")
    parser.add_argument("--filter-config", default=default_filter_config, help="Path to the sciobase filter config JSON.")
    parser.add_argument("--vcs", help="Override the VCS profile name from the filter config.")
    parser.add_argument("-o", "--output", help="Path to output processed JSON.")
    args = parser.parse_args()

    if not os.path.exists(args.input_json):
        print(f"Error: Input file {args.input_json} not found.")
        sys.exit(1)

    input_data = load_json_file(args.input_json)
    vcs_label = args.vcs or detect_vcs_label(input_data)
    selected_vcs, filter_profile = load_filter_config(args.filter_config, vcs_label)
    apply_filter_profile(filter_profile)

    print(f"Using filter profile: {selected_vcs}")

    comment_map = load_variant_comments(args.ref_dir)
    results = process_annotation_file(args.input_json, comment_map, filter_profile)

    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2)
        print(f"Processed variants saved to {args.output}")

        # Automatically generate Excel file alongside the JSON
        if args.output.endswith('.json'):
            xlsx_path = args.output[:-5] + '.xlsx'
        elif args.output.endswith('.json.gz'):
            xlsx_path = args.output[:-8] + '.xlsx'
        else:
            xlsx_path = args.output + '.xlsx'
            
        write_variants_to_xlsx(results, xlsx_path)
    else:
        print(json.dumps(results, indent=2))

if __name__ == "__main__":
    main()
